# -*- coding: utf-8 -*-
import re
from openpyxl import load_workbook, Workbook
from collections import defaultdict

rows_to_pass = set()
card_numbers = []


def fix_card_numbers(sheet, key):
    for i, col in enumerate(sheet[key],start=1):
        if i == 1: continue
        num = str(col.value)
        card_numbers.append(num)
        if not re.match('^\d+$', num):
            print(i, num)
            rows_to_pass.add(i)


def make_phone(phone, i):
    if phone is None:
        return 'NULL'
    phone = re.sub('[^\d]', '', str(phone))
    if phone == '' or phone == 'NULL':
        return 'NULL'
    if re.match('^[^78]\d{9}$', phone):
        phone = '7{}'.format(phone)
        print('row: {} 10 digit number + 7: {}'.format(i, phone))
    if re.match('^8\d{10}$', phone):
        phone = '7' + phone[1:]
        print('row: {} replace 8 with 7: {}'.format(i, phone))
    if len(phone) != 11:
        return 'NULL'
    return phone


def fix_phones(sheet, key1, key2):
    phones = defaultdict(int)
    for i, (col, col2) in enumerate(zip(sheet[key1], sheet[key2]), start=1):
        if i == 1 or i in rows_to_pass: continue
        phone = make_phone(col.value, i)
        if phone == 'NULL':
            phone2 = make_phone(col2.value, i)
            if phone2 != 'NULL':
                phone = phone2
            else:
                col.value = 'NULL'
                continue
        phones[phone] += 1
        col.value = int(phone)
    return phones


def fix_emails(sheet, key):
    emails = defaultdict(int)
    for i, col in enumerate(sheet[key], start=1):
        if i == 1 or i in rows_to_pass: continue
        email = col.value
        if email is None or type(email) == int or '@' not in email:
            col.value = 'NULL'
            continue
        email = re.sub('\s', '', email)
        awoid_domains = ['bj-gold.ru', 'bronnitsy.com', 'noemail.ru', 'no@mail.ru']
        if any(domain in email for domain in awoid_domains):
            print(i)
            rows_to_pass.add(i)
        emails[email] += 1
        col.value = email
    return emails


def find_equal(sheet, key, items):
    repeated_items = {k: value for k, value in items.items() if value > 1}
    repeated_rows = defaultdict(list)
    for i, col in enumerate(sheet[key], start=1):
        if i == 1 or i in rows_to_pass: continue
        item = str(col.value)
        if item in repeated_items:
            repeated_rows[item].append(i)
    for k in repeated_rows:
        print(k)
        repeated_rows[k].sort(key=lambda index: card_numbers[index])
    return repeated_rows


def merge_data(sheet, repeated_data):
    """repeated_data: dict with key of repeated instance(email or phone)
    and list of rows where the key is repeated"""
    # заменяем более современными значениями
    # ячейки в расположены по возрастанию, поэтому будет идти от самой последней к самой первой
    # и заполнять пропуски, если они имеются
    for key, rows in repeated_data.items():
        data = {}
        for row in rows[::-1]:
            for cell in sheet[row]:
                if cell.column not in data or data[cell.column] == 'NULL':
                    data[cell.column] = cell.value
                cell.value = None
        for cell in sheet[rows[-1]]:
            cell.value = data[cell.column]
        for row in rows[:-1]:
            rows_to_pass.add(row)
        print(data)


if __name__ == '__main__':
    print('Loading file started')
    original = load_workbook('main.xlsx')
    print('Loading file finished')
    fix_card_numbers(original['cols-2-upload'], 'A')
    print('Fix phones started')
    phones_1 = fix_phones(original['cols-2-upload'], 'E', 'F')
    print('Fix phones finished')
    print('Fix emails started')

    emails = fix_emails(original['cols-2-upload'], 'D')
    print('Fix emails finished')
    print('Find equal phones')
    phone_1_cells = find_equal(original['cols-2-upload'], 'E', phones_1)
    print('Find equal emails')
    email_cells = find_equal(original['cols-2-upload'], 'D', emails)
    print('Merge phones')
    merge_data(original['cols-2-upload'], phone_1_cells)
    print('Merge emails')
    merge_data(original['cols-2-upload'], email_cells)

    fix_card_numbers(original['cols-2-upload'], 'A')
    print('Start saving')
    new_wb = Workbook()
    ws = new_wb.active
    ws.title = 'cols-2-upload'
    rs = [r for i, r in enumerate(original['cols-2-upload'].rows, start=1) if i not in rows_to_pass]
    for i, r in enumerate(rs, start=1):
        for j, c in enumerate(r, start=1):
            ws.cell(row=i, column=j, value=c.value)
    new_wb.save('updated_1.xlsx')


# import threading
# from functools import partial
# from concurrent.futures import ThreadPoolExecutor
# def process(tupl, sheet):
#     key, rows = tupl
#     data = {}
#     for row in rows[::-1]:
#         for cell in sheet[row]:
#             if cell.column not in data or data[cell.column] == 'NULL':
#                 data[cell.column] = cell.value
#     with :
#         for cell in sheet[rows[-1]]:
#             cell.value = data[cell.column]
#         for row in rows[:-1]:
#             rows_to_pass.add(row)
#     print(data)
#
#
# def merge_data(sheet, repeated_data):
#     """repeated_data: dict with key of repeated instance(email or phone)
#     and list of rows where the key is repeated"""
#     # заменяем более современными значениями
#     # ячейки в расположены по возрастанию, поэтому будет идти от самой последней к самой первой
#     # и заполнять пропуски, если они имеются
#     extended_map = partial(process, sheet=sheet)
#     args = list(repeated_data.items())
#     with ThreadPoolExecutor(max_workers=5) as tpe:
#         for _ in tpe.map(extended_map, args):
#             pass